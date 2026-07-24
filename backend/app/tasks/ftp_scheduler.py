"""
ftp_scheduler.py
APScheduler 定时任务：每5分钟检查所有启用的 OSAT，
判断当前时间是否在抓取窗口内，若是则在线程池中异步执行抓取。
"""
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
from apscheduler.schedulers.background import BackgroundScheduler

# 线程池：最多同时抓取 4 个 OSAT（避免资源耗尽）
_executor = ThreadPoolExecutor(max_workers=4, thread_name_prefix="ftp_worker")
_scheduler = BackgroundScheduler(timezone="Asia/Shanghai")


# ──────────────────────────────────────────
# 时间窗口判断（支持跨午夜）
# ──────────────────────────────────────────

def is_in_window(start_str: str, end_str: str) -> bool:
    """
    判断当前时间是否在 [start_str, end_str] 窗口内。
    支持跨午夜：22:00~08:00 → now>=22:00 OR now<08:00
    """
    from zoneinfo import ZoneInfo
    now = datetime.now(ZoneInfo('Asia/Shanghai'))
    now_min = now.hour * 60 + now.minute

    sh, sm = start_str.split(':')
    eh, em = end_str.split(':')
    start_min = int(sh) * 60 + int(sm)
    end_min = int(eh) * 60 + int(em)

    if start_min <= end_min:
        # 同一天：08:00~22:00
        return start_min <= now_min < end_min
    else:
        # 跨午夜：22:00~08:00
        return now_min >= start_min or now_min < end_min


# ──────────────────────────────────────────
# 主调度任务（每5分钟触发）
# ──────────────────────────────────────────

def ftp_check_job():
    """
    每5分钟检查所有启用的 OSAT，
    对处于抓取窗口内的 OSAT 提交后台线程执行抓取。
    """
    from app.core.database import SessionLocal
    from app.models.osat_config import OsatConfig
    from app.services.ftp_service import run_osat_fetch

    db = SessionLocal()
    try:
        osats = db.query(OsatConfig).filter(OsatConfig.enabled == True).all()
        for osat in osats:
            if is_in_window(osat.schedule_start, osat.schedule_end):
                print(f"[scheduler] 触发 OSAT={osat.name} 抓取任务")
                _executor.submit(run_osat_fetch, osat.id, False)
            else:
                pass  # 不在时间窗口，静默跳过
    except Exception as e:
        print(f"[scheduler] ftp_check_job 异常: {e}")
    finally:
        db.close()


# ──────────────────────────────────────────
# 启动 / 停止 调度器
# ──────────────────────────────────────────

def send_daily_failure_report_job():
    """
    每天 09:00 整生成失败文件汇总的 Excel 报表，并发送给所有已订阅告警的 admin/eng 用户。
    """
    from app.core.database import SessionLocal
    from app.models.ftp_upload_log import FtpUploadLog
    from app.models.osat_config import OsatConfig
    from app.models.user import User
    from app.services.smtp_dynamic import send_smtp_attachment_auto
    from sqlalchemy import func
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    from collections import defaultdict
    
    db = SessionLocal()
    try:
        # 1. 查出所有当前仍处于失败状态的文件（排除已成功过的）
        success_paths = db.query(FtpUploadLog.remote_path).filter(FtpUploadLog.status == 'success').distinct()
        success_path_set = {r[0] for r in success_paths.all()}

        # 失败次数达到或超过 1 次（包含卡住的），取每个路径最新的失败记录
        subq = db.query(
            FtpUploadLog.remote_path,
            func.max(FtpUploadLog.id).label('max_id')
        ).filter(
            FtpUploadLog.status == 'failed'
        ).group_by(FtpUploadLog.remote_path).subquery()

        failed_logs = db.query(FtpUploadLog).join(
            subq, FtpUploadLog.id == subq.c.max_id
        ).all()

        failed_logs = [log for log in failed_logs if log.remote_path not in success_path_set]

        # 2. 如果没有任何失败记录，直接静默退出
        if not failed_logs:
            print("[scheduler] 每日失效报告生成：无失败文件，跳过发送")
            return

        # 3. 统计每个路径的重试失败次数
        fail_counts = db.query(
            FtpUploadLog.remote_path,
            func.count(FtpUploadLog.id).label('cnt')
        ).filter(FtpUploadLog.status == 'failed').group_by(FtpUploadLog.remote_path).all()
        fail_count_map = {r[0]: r[1] for r in fail_counts}

        # 4. 获取 OSAT 信息映射
        osats = db.query(OsatConfig).all()
        osat_map = {o.id: o for o in osats}
        osat_names = {o.id: o.name for o in osats}

        # 5. 各个 OSAT 失效情况分类汇总数据
        osat_summary = defaultdict(lambda: {
            'summary_count': 0,
            'summary_errors': defaultdict(int),
            'data_count': 0,
            'data_errors': defaultdict(int)
        })

        for log in failed_logs:
            osat_name = osat_names.get(log.osat_id, f"Unknown")
            is_summary = (log.filename or "").lower().endswith('.txt') or (log.remote_path or "").lower().endswith('.txt')
            err = log.error_msg or "未知错误"
            
            stats = osat_summary[osat_name]
            if is_summary:
                stats['summary_count'] += 1
                stats['summary_errors'][err] += 1
            else:
                stats['data_count'] += 1
                stats['data_errors'][err] += 1

        # 6. 开始生成 Excel
        wb = openpyxl.Workbook()
        
        # 第一页：各个OSAT失效汇总
        ws1 = wb.active
        ws1.title = "各个OSAT失效汇总"
        
        # 第二页：失败详细列表
        ws2 = wb.create_sheet(title="失败详细列表")
        
        # 格式样式设置
        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell_font = Font(name="微软雅黑", size=10)
        border_side = Side(border_style="thin", color="D3D3D3")
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        
        # 填充第一页表头
        ws1.append(["OSAT 厂商名称", "Summary失效文件数", "Summary失效原因分布", "数据失效文件数", "数据失效原因分布"])
        for cell in ws1[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            
        for osat_name, stats in osat_summary.items():
            sum_err_str = ", ".join(f"{k}({v}个)" for k, v in stats['summary_errors'].items())
            dat_err_str = ", ".join(f"{k}({v}个)" for k, v in stats['data_errors'].items())
            ws1.append([
                osat_name,
                stats['summary_count'],
                sum_err_str or "无",
                stats['data_count'],
                dat_err_str or "无"
            ])
            
        # 填充第二页表头
        ws2.append(["OSAT 厂商", "文件名", "FTP 远程路径", "FTP 链接地址 (可用于客户端)", "失败原因", "尝试失败次数", "最后尝试时间"])
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            
        for log in failed_logs:
            osat_name = osat_names.get(log.osat_id, f"Unknown")
            osat_cfg = osat_map.get(log.osat_id)
            if osat_cfg:
                proto = osat_cfg.protocol or "ftp"
                port = osat_cfg.ftp_port or (22 if proto == "sftp" else 21)
                ftp_url = f"{proto}://{osat_cfg.ftp_user}@{osat_cfg.ftp_host}:{port}{log.remote_path}"
            else:
                ftp_url = log.remote_path
                
            fail_cnt = fail_count_map.get(log.remote_path, 1)
            last_attempt = log.uploaded_at.strftime("%Y-%m-%d %H:%M:%S") if log.uploaded_at else "—"
            
            ws2.append([
                osat_name,
                log.filename or "—",
                log.remote_path,
                ftp_url,
                log.error_msg or "未知错误",
                fail_cnt,
                last_attempt
            ])
            
        # 设置框线和自适应列宽
        for ws in (ws1, ws2):
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = cell_font
                    cell.border = border
                    if isinstance(cell.value, int):
                        cell.alignment = align_center
                    else:
                        cell.alignment = align_left
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 65)
                
        # 写入 BytesIO 缓冲区
        out_buf = BytesIO()
        wb.save(out_buf)
        excel_data = out_buf.getvalue()
        wb.close()
        
        # 7. 获取要发送的接收人邮箱（有订阅的 admin / eng 用户）
        alert_users = db.query(User).filter(
            User.receive_alerts == True,
            User.is_active == True,
            User.role.in_(["admin", "eng"]),
            User.email.isnot(None)
        ).all()
        
        emails = [u.email for u in alert_users]
        if not emails:
            # 备用：发给第一个 role = 'admin' 用户
            admin = db.query(User).filter(User.role == 'admin').first()
            if admin and admin.email:
                emails = [admin.email]
                
        if not emails:
            print("[scheduler] 每日失效报告生成：未找到任何接收邮箱")
            return
            
        # 8. 邮件发送
        from datetime import datetime
        report_date = datetime.now().strftime("%Y-%m-%d")
        subject = f"【ATE系统】每日失效文件汇总报告 - {report_date}"
        html_body = f"""
        <div style="font-family:sans-serif;max-width:600px;margin:auto;border:1px solid #ddd;padding:20px;border-radius:8px">
          <h2 style="color:#2f3640;border-bottom:2px solid #356092;padding-bottom:10px">📋 ATE 系统每日失效文件汇总报告</h2>
          <p>您好！这是系统自动生成的每日失效文件汇总报告。</p>
          <p>报告日期：<b>{report_date} 09:00</b></p>
          <p>当前系统内共有 <b>{len(failed_logs)}</b> 个文件处于失效/重试超限状态。详细的失效分类汇总与具体 FTP 链接信息已整理在随信附件的 Excel 表格中，请查收并进行排查与手动上传确认。</p>
          <hr style="border:0;border-top:1px solid #eee;margin:20px 0"/>
          <p style="color:#7f8c8d;font-size:12px">此邮件为系统自动触发发送，请勿直接回复。</p>
        </div>
        """
        attachment_name = f"ATE_Failed_Files_Report_{report_date}.xlsx"
        
        for email in emails:
            try:
                send_smtp_attachment_auto(
                    db=db,
                    to_email=email,
                    subject=subject,
                    html_body=html_body,
                    attachment_bytes=excel_data,
                    attachment_name=attachment_name
                )
                print(f"[scheduler] 每日失效报告已成功发送至：{email}")
            except Exception as ex:
                print(f"[scheduler] 发送每日失效报告至 {email} 失败: {ex}")
                
    except Exception as e:
        print(f"[scheduler] send_daily_failure_report_job 异常: {e}")
    finally:
        db.close()


def precalculate_mp_yield_job():
    """Precalculate the 12 combinations of mass production yield snapshots daily at 02:00:00 and save to Redis, TTL = 8 days"""
    from app.core.database import SessionLocal
    from app.api.routes.lots import get_mp_yield_overview
    import time

    print("[scheduler] ⏰ Starting daily mass production yield overview precalculation job...")
    
    combinations = [
        ("month", 1), ("month", 3), ("month", 6), ("month", 12),
        ("year", 1), ("year", 2), ("year", 3),
        ("lot", 20), ("lot", 50), ("lot", 100), ("lot", 200),
        ("all", 0)
    ]
    
    db = SessionLocal()
    try:
        for r_type, r_val in combinations:
            t0 = time.time()
            try:
                get_mp_yield_overview(
                    db=db,
                    current_user=None,
                    range_type=r_type,
                    range_value=r_val if r_type != "all" else None,
                    months=None,
                    product_name=None,
                    bypass_cache=True
                )
                print(f"[scheduler] Successfully computed and cached combination ({r_type}, {r_val}), time: {time.time() - t0:.2f} seconds")
            except Exception as ex:
                print(f"[scheduler] Precalculating combination ({r_type}, {r_val}) failed: {ex}")
        print("[scheduler] ✅ Daily mass production yield overview precalculation job completed successfully!")
    except Exception as e:
        print(f"[scheduler] precalculate_mp_yield_job exception: {e}")
    finally:
        db.close()


# Directory Growth Watchdog State
_last_dir_sizes = {
    "/tmp/FTP/download": 0,
    "/tmp/FTP/extracted": 0,
    "/app/uploads": 0,
    "/app": 0
}

def get_dir_size(path):
    import os
    total = 0
    if not os.path.exists(path):
        return 0
    try:
        for root, _, files in os.walk(path):
            for f in files:
                fp = os.path.join(root, f)
                try:
                    total += os.path.getsize(fp)
                except:
                    pass
    except Exception:
        pass
    return total

def directory_growth_monitor_job():
    '''
    Directory growth watchdog with auto-resumption support.
    Runs every 1 minute to check size changes of:
    - /tmp/FTP/download
    - /tmp/FTP/extracted
    - /app/uploads
    - /app
    '''
    global _last_dir_sizes
    import shutil
    import json
    from app.core.database import SessionLocal
    from app.models.osat_config import OsatConfig
    from app.services.smtp_dynamic import send_smtp_auto
    from app.models.user import User
    from app.core.redis_client import get_redis

    dirs = ["/tmp/FTP/download", "/tmp/FTP/extracted", "/app/uploads", "/app"]
    limit_increase = 4 * 1024 * 1024 * 1024  # 4 GB
    limit_absolute = 40 * 1024 * 1024 * 1024  # 40 GB absolute limit for temp folders
    resume_threshold = 3 * 1024 * 1024 * 1024  # 3 GB resume threshold

    triggered = False
    trigger_reason = ""

    # 1. Check directory sizes and growth rates
    temp_total_size = 0
    for d in dirs:
        current_size = get_dir_size(d)
        last_size = _last_dir_sizes.get(d, 0)
        
        # Track total size of temp folders
        if d in ["/tmp/FTP/download", "/tmp/FTP/extracted"]:
            temp_total_size += current_size

        # Initialize if it was 0 (first run)
        if last_size == 0 and current_size > 0:
            _last_dir_sizes[d] = current_size
            continue

        increase = current_size - last_size
        _last_dir_sizes[d] = current_size

        if increase > limit_increase:
            triggered = True
            trigger_reason = f"Directory {d} grew by {increase / (1024**2):.1f} MB (exceeded 4GB limit) in 1 minute."
            break
            
        # Absolute limit check for temp folders
        if d in ["/tmp/FTP/download", "/tmp/FTP/extracted"] and current_size > limit_absolute:
            triggered = True
            trigger_reason = f"Directory {d} size is {current_size / (1024**3):.1f} GB (exceeded 40GB absolute limit)."
            break

    # 2. Check partition free space as double-insurance
    if not triggered:
        try:
            _, _, free_bytes = shutil.disk_usage("/app")
            if free_bytes < 5 * 1024 * 1024 * 1024:  # 5 GB free space threshold
                triggered = True
                trigger_reason = f"Host partition free space is dangerously low: {free_bytes / (1024**3):.1f} GB remaining."
        except Exception as d_ex:
            print(f"[watchdog] Failed to check disk usage: {d_ex}")

    if triggered:
        print(f"[watchdog] ??? CRITICAL: {trigger_reason} Performing emergency stop!")
        db = SessionLocal()
        try:
            # 1. Save enabled OSAT IDs and disable them
            osats = db.query(OsatConfig).filter(OsatConfig.enabled == True).all()
            forced_ids = [o.id for o in osats]
            
            if forced_ids:
                try:
                    r = get_redis()
                    r.set("watchdog:forced_disabled_osats", json.dumps(forced_ids))
                    print(f"[watchdog] Saved forced disabled OSAT IDs: {forced_ids}")
                except Exception as redis_ex:
                    print(f"[watchdog] Failed to save forced disabled OSAT IDs to Redis: {redis_ex}")

                for o in osats:
                    o.enabled = False
                    print(f"[watchdog] OSAT={o.name} has been auto-disabled.")
                db.commit()

            # 2. Rate-limit email alerts
            should_send_email = False
            try:
                import time
                from datetime import datetime
                from zoneinfo import ZoneInfo
                
                r = get_redis()
                tz = ZoneInfo("Asia/Shanghai")
                now_dt = datetime.now(tz)
                now_ts = time.time()
                
                is_first_alert = not r.get("watchdog:alert_state")
                
                if is_first_alert:
                    should_send_email = True
                    r.set("watchdog:alert_state", "triggered")
                    r.set("watchdog:last_alert_time", str(now_ts))
                else:
                    is_weekday = now_dt.weekday() < 5  # Mon-Fri
                    is_working_hours = 8 <= now_dt.hour < 20
                    
                    if is_weekday and is_working_hours:
                        last_alert_str = r.get("watchdog:last_alert_time")
                        last_alert_ts = float(last_alert_str) if last_alert_str else 0.0
                        if now_ts - last_alert_ts >= 4 * 3600:
                            should_send_email = True
                            r.set("watchdog:last_alert_time", str(now_ts))
                    else:
                        # Outside working hours / weekends: do not send subsequent alerts
                        pass
            except Exception as redis_ex:
                print(f"[watchdog] Throttling logic failed (fallback to always send): {redis_ex}")
                should_send_email = True

            if should_send_email:
                admins = db.query(User).filter(User.role == 'admin', User.is_active == True).all()
                recipient_emails = [a.email for a in admins if a.email]
                if recipient_emails:
                    subject = "??RITICAL ALERT??TE Server Disk Space growth Emergency Stop!"
                    body_content = (
                        f"Warning: The ATE Server directory growth watchdog has triggered an emergency shutdown.\n\n"
                        f"Reason:\n{trigger_reason}\n\n"
                        f"Action Taken:\nAll FTP OSAT schedulers have been automatically set to DISABLED to prevent server disk overflow.\n\n"
                        f"Please check the server directories and logs immediately."
                    )
                    for email in recipient_emails:
                        try:
                            send_smtp_auto(db, email, subject, body_content)
                            print(f"[watchdog] SMTP Alert sent to {email}")
                        except Exception as s_ex:
                            print(f"[watchdog] Failed to send SMTP alert to {email}: {s_ex}")
            else:
                print("[watchdog] Alert email throttled, skipping email send.")
        except Exception as ex:
            print(f"[watchdog] Error executing emergency stop: {ex}")
            db.rollback()
        finally:
            db.close()
    else:
        # Healthy state:
        # 1. Clear alert Redis keys
        try:
            r = get_redis()
            if r.get("watchdog:alert_state"):
                r.delete("watchdog:alert_state")
                r.delete("watchdog:last_alert_time")
                print("[watchdog] System recovered to healthy state, cleared alert throttling state.")
        except Exception:
            pass

        # 2. Check if we should auto-resume OSATs (total temp size < 1GB)
        if temp_total_size < resume_threshold:
            try:
                r = get_redis()
                forced_ids_str = r.get("watchdog:forced_disabled_osats")
                if forced_ids_str:
                    forced_ids = json.loads(forced_ids_str)
                    if forced_ids:
                        db = SessionLocal()
                        try:
                            osats = db.query(OsatConfig).filter(OsatConfig.id.in_(forced_ids)).all()
                            for o in osats:
                                o.enabled = True
                                print(f"[watchdog] OSAT={o.name} has been auto-enabled (temp size = {temp_total_size / (1024**2):.1f} MB).")
                            db.commit()
                            r.delete("watchdog:forced_disabled_osats")
                        finally:
                            db.close()
            except Exception as resume_ex:
                print(f"[watchdog] Error trying to auto-resume OSATs: {resume_ex}")


def daily_ftp_scan_snapshot_job():
    """Run daily scan snapshot at 08:00 for all enabled OSATs."""
    from app.core.database import SessionLocal
    from app.models.osat_config import OsatConfig
    from app.services.ftp_service import run_osat_fetch
    
    print("[scheduler] ⏰ Starting daily FTP scan snapshot job (08:00)...")
    db = SessionLocal()
    try:
        configs = db.query(OsatConfig).filter(OsatConfig.enabled == True).all()
        for o in configs:
            _executor.submit(run_osat_fetch, o.id, True)
    except Exception as e:
        print(f"[scheduler] daily_ftp_scan_snapshot_job exception: {e}")
    finally:
        db.close()


def cleanup_del_folder_job():
    """Clean up the '/tmp/FTP/del' folder once every 5 hours."""
    import os, shutil
    del_dir = "/tmp/FTP/del"
    if os.path.exists(del_dir):
        print(f"[scheduler] Starting cleanup of {del_dir}...")
        for name in os.listdir(del_dir):
            path = os.path.join(del_dir, name)
            try:
                if os.path.isdir(path):
                    shutil.rmtree(path)
                else:
                    os.remove(path)
            except Exception as e:
                print(f"[scheduler] Failed to delete {path}: {e}")
        print("[scheduler] Cleanup of del folder completed.")


def start_scheduler():
    """在应用启动时调用，注册定时任务"""
    if not _scheduler.running:
        from zoneinfo import ZoneInfo
        shanghai_tz = ZoneInfo("Asia/Shanghai")
        _scheduler.add_job(
            ftp_check_job,
            trigger='interval',
            minutes=5,
            id='ftp_check',
            next_run_time=datetime.now(shanghai_tz),
            replace_existing=True,
            misfire_grace_time=60,
        )
        # Clean up '/tmp/FTP/del' folder every day at 18:00 (24h cycle)
        _scheduler.add_job(
            cleanup_del_folder_job,
            trigger='cron',
            hour=18,
            minute=0,
            id='cleanup_del_folder',
            replace_existing=True,
            misfire_grace_time=3600,
        )
        # 每 1 分钟检查一次本地临时目录的增长速度，防止异常爆盘
        _scheduler.add_job(
            directory_growth_monitor_job,
            trigger='interval',
            minutes=1,
            id='directory_growth_monitor',
            replace_existing=True,
            misfire_grace_time=10,
        )
        # 每天 08:00 整执行 FTP 扫描并更新快照
        _scheduler.add_job(
            daily_ftp_scan_snapshot_job,
            trigger='cron',
            hour=8,
            minute=0,
            id='daily_ftp_scan_snapshot',
            replace_existing=True,
            misfire_grace_time=3600,
        )
        # 每天 09:00 整将失效文件汇总并发件
        _scheduler.add_job(
            send_daily_failure_report_job,
            trigger='cron',
            hour=9,
            minute=0,
            id='daily_failure_report',
            replace_existing=True,
            misfire_grace_time=3600,
        )
        # Daily 02:00:00 precalculate mass production yield overview snapshot
        _scheduler.add_job(
            precalculate_mp_yield_job,
            trigger='cron',
            hour=2,
            minute=0,
            id='precalculate_mp_yield_snapshot',
            replace_existing=True,
            misfire_grace_time=3600,
        )
        _scheduler.start()
        print("[scheduler] FTP 定时调度器已启动（包含每5分钟检查、每日08:00扫描及每日09:00汇总报告）")


def stop_scheduler():
    """在应用关闭时调用"""
    if _scheduler.running:
        _scheduler.shutdown(wait=False)
        _executor.shutdown(wait=False)
        print("[scheduler] FTP 定时调度器已停止")


def trigger_osat_now(osat_id: int):
    """手动立即触发一个 OSAT 的抓取任务（在后台线程中执行）"""
    from app.services.ftp_service import run_osat_fetch
    _executor.submit(run_osat_fetch, osat_id, True)
    print(f"[scheduler] 手动触发 OSAT id={osat_id} 抓取任务")
