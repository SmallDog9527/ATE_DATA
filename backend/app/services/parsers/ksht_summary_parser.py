import os
import re
import openpyxl
from datetime import datetime
from sqlalchemy.orm import Session
from app.models.lot import Lot, DataSource, ProcessStatus
from app.models.bin_summary import BinSummary

def parse_ksht_date(val):
    if not val:
        return None
    s = str(val).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None

def parse_and_save_ksht_summary(filepath: str, db: Session, user_id: int = None, osat_name: str = "HTKS") -> list:
    """
    Parse KSHT XLSX summary reports (both Format 1 and Format 2)
    and save the parsed wafer data into the database.
    """
    filename = os.path.basename(filepath)
    print(f"[ksht_summary_parser] Parsing file: {filename}")
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_names = wb.sheetnames
    
    parsed_wafers = []
    
    if "Lot" in sheet_names:
        # ── Parse Format 1 ───────────────────────────────────────────────────
        print("[ksht_summary_parser] Detected Format 1 (with 'Lot' sheet)")
        sh_lot = wb["Lot"]
        device_name = ""
        for r in range(1, sh_lot.max_row + 1):
            cell_k = sh_lot.cell(r, 1).value
            cell_v = sh_lot.cell(r, 2).value
            if cell_k:
                k_str = str(cell_k).strip().lower()
                if k_str == "customer device":
                    device_name = str(cell_v).strip()
                    
        sh_sum = wb["Summary"]
        row1 = [sh_sum.cell(1, c).value for c in range(1, sh_sum.max_column + 1)]
        row2 = [sh_sum.cell(2, c).value for c in range(1, sh_sum.max_column + 1)]
        
        try:
            c_wafer = row1.index("Wafer")
        except ValueError:
            c_wafer = 0
            
        c_tester = row1.index("Tester") if "Tester" in row1 else -1
        c_probercard = row1.index("ProberCard") if "ProberCard" in row1 else -1
        c_started = row1.index("Started") if "Started" in row1 else -1
        c_finished = row1.index("Finished") if "Finished" in row1 else -1
        c_job = row1.index("Job") if "Job" in row1 else -1
        c_gross = row1.index("Gross") if "Gross" in row1 else -1
        
        # Find Qty start column index in row1/row2
        c_qty_start = row1.index("Qty") if "Qty" in row1 else -1
        if c_qty_start < 0:
            try:
                first_idx = row2.index("Good 1st")
                c_qty_start = row2.index("Good 1st", first_idx + 1)
            except ValueError:
                c_qty_start = -1
                
        # Bin mapping from Qty columns in row 2
        bin_cols = {}
        if c_qty_start >= 0:
            for c in range(c_qty_start + 3, len(row2)):
                val = row2[c]
                if val is not None:
                    val_str = str(val).strip()
                    if val_str.isdigit():
                        bin_cols[c] = int(val_str)
                        
        # Iterate data rows starting from row 3
        for r in range(3, sh_sum.max_row + 1):
            wafer_key = sh_sum.cell(r, c_wafer + 1).value
            if not wafer_key or str(wafer_key).strip().lower() in ("total", "average", "percentage"):
                continue
            
            wafer_key_str = str(wafer_key).strip()
            m = re.match(r"^([^-]+)-(\d+)", wafer_key_str)
            if not m:
                continue
            lot_id = m.group(1).strip()
            wafer_id = str(int(m.group(2).strip()))
            
            mp_tester = sh_sum.cell(r, c_tester + 1).value if c_tester >= 0 else None
            probecard = sh_sum.cell(r, c_probercard + 1).value if c_probercard >= 0 else None
            test_start = parse_ksht_date(sh_sum.cell(r, c_started + 1).value) if c_started >= 0 else None
            test_finish = parse_ksht_date(sh_sum.cell(r, c_finished + 1).value) if c_finished >= 0 else None
            program = sh_sum.cell(r, c_job + 1).value if c_job >= 0 else None
            die_count = sh_sum.cell(r, c_gross + 1).value
            
            yield_val = sh_sum.cell(r, 10).value # Good rate at Column 10 (index 9)
            yield_rate = None
            if yield_val is not None:
                yield_rate = float(yield_val)
                
            if die_count is not None:
                die_count = int(die_count)
            
            pass_count = None
            if c_qty_start >= 0:
                pass_val = sh_sum.cell(r, c_qty_start + 2).value
                if pass_val is not None:
                    pass_count = int(pass_val)
            
            if pass_count is None and die_count and yield_rate is not None:
                pass_count = int(round(die_count * yield_rate))
                
            sbin_counts = {}
            if pass_count is not None:
                sbin_counts[1] = pass_count
                
            for c_idx, b_num in bin_cols.items():
                qty_val = sh_sum.cell(r, c_idx + 1).value
                if qty_val is not None:
                    sbin_counts[b_num] = int(qty_val)
            
            parsed_wafers.append({
                "lot_id": lot_id,
                "wafer_id": wafer_id,
                "device_name": device_name,
                "mp_tester": mp_tester,
                "probecard": probecard,
                "test_start": test_start,
                "test_finish": test_finish,
                "program": program,
                "die_count": die_count,
                "pass_count": pass_count,
                "yield_rate": yield_rate,
                "sbin_counts": sbin_counts
            })
            
    else:
        # ── Parse Format 2 ───────────────────────────────────────────────────
        print("[ksht_summary_parser] Detected Format 2 (with 'Summary' sheet only)")
        sh_sum = wb["Summary"]
        header_row_idx = None
        for r in range(1, sh_sum.max_row + 1):
            val = sh_sum.cell(r, 1).value
            if val and str(val).strip().lower() == "test start time":
                header_row_idx = r
                break
                
        if header_row_idx is None:
            raise ValueError("Could not find 'Test Start Time' header row in Summary sheet")
            
        row_headers = [sh_sum.cell(header_row_idx, c).value for c in range(1, sh_sum.max_column + 1)]
        
        c_started = row_headers.index("Test Start Time") if "Test Start Time" in row_headers else -1
        c_finished = row_headers.index("Test Finish Time") if "Test Finish Time" in row_headers else -1
        c_device = row_headers.index("Device Name") if "Device Name" in row_headers else -1
        c_wafer = row_headers.index("WaferID") if "WaferID" in row_headers else -1
        c_total = row_headers.index("Total") if "Total" in row_headers else -1
        c_pass = row_headers.index("Pass") if "Pass" in row_headers else -1
        c_yield = row_headers.index("Yield") if "Yield" in row_headers else -1
        
        bin_cols = {}
        for c in range(len(row_headers)):
            val = row_headers[c]
            if val:
                val_str = str(val).strip().upper()
                if val_str.startswith("BIN "):
                    try:
                        b_num = int(val_str.split()[1])
                        bin_cols[c] = b_num
                    except Exception:
                        pass
                        
        for r in range(header_row_idx + 1, sh_sum.max_row + 1):
            wafer_key = sh_sum.cell(r, c_wafer + 1).value
            if not wafer_key or str(wafer_key).strip().lower() in ("total", "average", "percentage"):
                continue
                
            wafer_key_str = str(wafer_key).strip()
            m = re.match(r"^([^-]+)-(\d+)", wafer_key_str)
            if not m:
                continue
            lot_id = m.group(1).strip()
            wafer_id = str(int(m.group(2).strip()))
            
            device_name = sh_sum.cell(r, c_device + 1).value if c_device >= 0 else ""
            test_start = parse_ksht_date(sh_sum.cell(r, c_started + 1).value) if c_started >= 0 else None
            test_finish = parse_ksht_date(sh_sum.cell(r, c_finished + 1).value) if c_finished >= 0 else None
            die_count = sh_sum.cell(r, c_total + 1).value
            pass_count = sh_sum.cell(r, c_pass + 1).value
            yield_val = sh_sum.cell(r, c_yield + 1).value
            
            if die_count is not None:
                die_count = int(die_count)
            if pass_count is not None:
                pass_count = int(pass_count)
                
            yield_rate = None
            if yield_val is not None:
                try:
                    yield_rate = float(str(yield_val).replace('%', '').strip()) / 100.0
                except ValueError:
                    pass
                    
            sbin_counts = {}
            if pass_count is not None:
                sbin_counts[1] = pass_count
                
            for c_idx, b_num in bin_cols.items():
                rate_val = sh_sum.cell(r, c_idx + 1).value
                if rate_val is not None and die_count:
                    try:
                        rate_float = float(str(rate_val).replace('%', '').strip()) / 100.0
                        sbin_counts[b_num] = int(round(die_count * rate_float))
                    except ValueError:
                        pass
                        
            parsed_wafers.append({
                "lot_id": lot_id,
                "wafer_id": wafer_id,
                "device_name": device_name,
                "mp_tester": None,
                "probecard": None,
                "test_start": test_start,
                "test_finish": test_finish,
                "program": None,
                "die_count": die_count,
                "pass_count": pass_count,
                "yield_rate": yield_rate,
                "sbin_counts": sbin_counts
            })
            
    # ── Save to Database ─────────────────────────────────────────────────────
    created_lots = []
    
    for w in parsed_wafers:
        lot_id = w["lot_id"]
        wafer_id = w["wafer_id"]
        
        # Check if lot with same lot_id, wafer_id and type already exists
        existing_lot = db.query(Lot).filter(
            Lot.lot_id == lot_id,
            Lot.wafer_id == wafer_id,
            Lot.data_type == 'MP_Yield'
        ).first()
        
        if existing_lot:
            # Overwrite: delete existing lot and its bin summary
            for bs in db.query(BinSummary).filter(BinSummary.lot_id == existing_lot.id).all():
                db.delete(bs)
            db.delete(existing_lot)
            db.commit()
            
        # Create new Lot record
        lot = Lot(
            filename=filename,
            storage_path=filepath,
            file_size=os.path.getsize(filepath),
            status=ProcessStatus.processed,
            data_source=DataSource.manual if user_id else DataSource.ftp,
            storage_type='local',
            upload_date=datetime.now(),
            user_id=user_id,
            
            product_name=w["device_name"],
            lot_id=lot_id,
            wafer_id=wafer_id,
            data_type='MP_Yield',
            test_stage='MP_Yield',
            test_machine=w["mp_tester"],
            mp_tester=w["mp_tester"],
            probecard=w["probecard"],
            program=w["program"],
            die_count=w["die_count"],
            pass_count=w["pass_count"],
            yield_rate=w["yield_rate"],
            osat_name=osat_name,
            test_date=w["test_finish"],
            beginning_time=w["test_start"],
            ending_time=w["test_finish"],
        )
        
        db.add(lot)
        db.commit()
        db.refresh(lot)
        
        # Save BinSummary records (1 to 130)
        for sbin_idx in range(1, 131):
            count = w["sbin_counts"].get(sbin_idx, 0)
            percentage = (count / w["die_count"]) * 100.0 if (w["die_count"] and w["die_count"] > 0) else 0.0
            
            bin_sum = BinSummary(
                lot_id=lot.id,
                bin_number=sbin_idx,
                bin_name=f"Sbin{sbin_idx}",
                site=0,
                count=count,
                percentage=percentage,
                data_range="final"
            )
            db.add(bin_sum)
            
        db.commit()
        created_lots.append(lot)
        
    print(f"[ksht_summary_parser] Successfully parsed and saved {len(created_lots)} lots from {filename}.")
    return created_lots
