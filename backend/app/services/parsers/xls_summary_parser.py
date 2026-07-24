import os
import openpyxl
import pandas as pd
from sqlalchemy.orm import Session

def parse_and_save_xls_summary(filepath: str, db: Session, user_id: int = None, osat_name: str = "Chipmore") -> list:
    """
    Unified entry point for parsing XLS/XLSX summary reports.
    Dispatches parsing requests to the correct OSAT parser based on the factory name,
    with auto-detection support for KSHT, LBS, and UCD formats.
    """
    sheet_names = []
    is_xls = filepath.lower().endswith('.xls')
    
    # 1. Determine sheet names using openpyxl (for xlsx) or xlrd (for xls)
    try:
        if is_xls:
            import xlrd
            wb = xlrd.open_workbook(filepath, on_demand=True)
            sheet_names = wb.sheet_names()
        else:
            wb = openpyxl.load_workbook(filepath, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
    except Exception as e:
        print(f"[xls_summary_parser] Error reading sheet names for auto-detection: {e}")

    # 1.5 Auto-detection check for "VENDOR" -> "CHIPMORE" in the first sheet
    is_vendor_chipmore = False
    try:
        if is_xls:
            import xlrd
            wb = xlrd.open_workbook(filepath)
            sh = wb.sheet_by_index(0)
            for r in range(min(15, sh.nrows)):
                for c in range(sh.ncols):
                    val = str(sh.cell_value(r, c)).strip().upper()
                    if val == "VENDOR":
                        for offset in (1, 2):
                            if c + offset < sh.ncols:
                                next_val = str(sh.cell_value(r, c + offset)).strip().upper()
                                if next_val == "CHIPMORE":
                                    is_vendor_chipmore = True
                                    break
                if is_vendor_chipmore:
                    break
        else:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            sh = wb.worksheets[0]
            max_r = min(15, sh.max_row or 15)
            max_c = min(30, sh.max_column or 30)
            for r in range(1, max_r + 1):
                for c in range(1, max_c + 1):
                    val = sh.cell(r, c).value
                    if val and str(val).strip().upper() == "VENDOR":
                        for offset in (1, 2):
                            if c + offset <= sh.max_column:
                                next_val = sh.cell(r, c + offset).value
                                if next_val and str(next_val).strip().upper() == "CHIPMORE":
                                    is_vendor_chipmore = True
                                    break
                if is_vendor_chipmore:
                    break
            wb.close()
    except Exception as e:
        print(f"[xls_summary_parser] Error checking VENDOR/CHIPMORE: {e}")

    # 1.6 Auto-detection check for UCD (first sheet, Row 3 Col 2 is Cust, Row 4 Col 2 is HMC)
    is_ucd = False
    try:
        if is_xls:
            import xlrd
            wb = xlrd.open_workbook(filepath)
            sh = wb.sheet_by_index(0)
            if sh.nrows >= 4 and sh.ncols >= 2:
                c2_header = str(sh.cell_value(2, 1)).strip().lower()
                c2_value = str(sh.cell_value(3, 1)).strip().upper()
                if c2_header == "cust" and c2_value == "HMC":
                    is_ucd = True
        else:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            sh = wb.worksheets[0]
            if sh.max_row >= 4 and sh.max_column >= 2:
                c2_header = sh.cell(3, 2).value
                c2_value = sh.cell(4, 2).value
                if c2_header and str(c2_header).strip().lower() == "cust":
                    if c2_value and str(c2_value).strip().upper() == "HMC":
                        is_ucd = True
            wb.close()
    except Exception as e:
        print(f"[xls_summary_parser] Error checking UCD: {e}")

    # 2. Determine parser name based on auto-detection or fallback to osat_name
    name = str(osat_name).strip().lower()
    if name == "htks":
        name = "ksht"
    
    
    if is_vendor_chipmore:
        print("[xls_summary_parser] Auto-detected Chipmore Format (has VENDOR: CHIPMORE)")
        name = "Chipmore"
    elif is_ucd:
        print("[xls_summary_parser] Auto-detected UCD Format (has Cust: HMC in first sheet)")
        name = "ucd"
    elif "Bin_Summary" in sheet_names:
        # Resolve Chipmore vs LBS based on column headers inside Bin_Summary sheet
        is_lbs = True
        filename_lower = os.path.basename(filepath).lower()
        if "lbs" in filename_lower or (osat_name and "lbs" in str(osat_name).lower()):
            is_lbs = True
        else:
            try:
                with pd.ExcelFile(filepath) as xl:
                    df = xl.parse("Bin_Summary", nrows=15, header=None)
                for r in range(df.shape[0]):
                    val = df.iloc[r, 0]
                    if pd.notna(val) and str(val).strip().lower() == "lotid-waferid":
                        if df.shape[1] > 1:
                            col1_val = str(df.iloc[r, 1]).strip().lower()
                            if col1_val == "yield(%)":
                                is_lbs = False
                                break
                            elif col1_val in ("total test", "total tested"):
                                is_lbs = True
                                break
            except Exception as e:
                print(f"[xls_summary_parser] Error checking Bin_Summary columns: {e}")
            
        if is_lbs:
            print("[xls_summary_parser] Auto-detected LBS Format (has 'Bin_Summary' sheet with LBS columns)")
            name = "lbs"
        else:
            print("[xls_summary_parser] Auto-detected Chipmore Format (has 'Bin_Summary' sheet with Chipmore columns)")
            name = "Chipmore"
    elif "Lot" in sheet_names:
        print("[xls_summary_parser] Auto-detected KSHT Format 1 (has 'Lot' sheet)")
        name = "ksht"
    elif "Summary" in sheet_names:
        is_ksht_format2 = False
        try:
            if is_xls:
                import xlrd
                wb = xlrd.open_workbook(filepath)
                sh = wb.sheet_by_name("Summary")
                for r in range(min(10, sh.nrows)):
                    row_vals = [str(sh.cell_value(r, c)).strip().lower() for c in range(sh.ncols)]
                    if any("test start time" in val for val in row_vals):
                        is_ksht_format2 = True
                        break
            else:
                wb = openpyxl.load_workbook(filepath, read_only=True)
                sh = wb["Summary"]
                for row in sh.iter_rows(max_row=10, values_only=True):
                    if any(x is not None for x in row):
                        row_vals = [str(x).strip().lower() for x in row if x is not None]
                        if any("test start time" in val for val in row_vals):
                            is_ksht_format2 = True
                            break
                wb.close()
        except Exception as e:
            print(f"[xls_summary_parser] Error checking KSHT format 2 headers: {e}")
        
        if is_ksht_format2:
            print("[xls_summary_parser] Auto-detected KSHT Format 2 (has 'Test Start Time' header)")
            name = "ksht"

    # Fallback to filename-based detection if sheet detection didn't resolve it
    if name not in ("lbs", "ksht", "chipmore", "Chipmore", "ucd"):
        filename_lower = os.path.basename(filepath).lower()
        if "lbs" in filename_lower:
            print("[xls_summary_parser] Auto-detected LBS via filename")
            name = "lbs"
        elif "ksht" in filename_lower or "htks" in filename_lower:
            print("[xls_summary_parser] Auto-detected KSHT via filename")
            name = "ksht"
        elif "ucd" in filename_lower:
            print("[xls_summary_parser] Auto-detected UCD via filename")
            name = "ucd"

    print(f"[xls_summary_parser] Routing summary report parse for OSAT: {name!r} (original: {osat_name!r})")
    
    if name == "ksht":
        from app.services.parsers.ksht_summary_parser import parse_and_save_ksht_summary
        return parse_and_save_ksht_summary(filepath, db, user_id, osat_name="HTKS")
    elif name == "lbs":
        from app.services.parsers.lbs_summary_parser import parse_and_save_lbs_summary
        return parse_and_save_lbs_summary(filepath, db, user_id, osat_name="LBS")
    elif name == "ucd":
        from app.services.parsers.ucd_summary_parser import parse_and_save_ucd_summary
        return parse_and_save_ucd_summary(filepath, db, user_id, osat_name="UCD")
    elif name in ("chipmore", "Chipmore") or not name:
        from app.services.parsers.chipmore_summary_parser import parse_and_save_chipmore_summary
        return parse_and_save_chipmore_summary(filepath, db, user_id, osat_name="Chipmore")
    else:
        print(f"[xls_summary_parser] Warning: Unknown OSAT {osat_name!r}. Falling back to Chipmore parser.")
        from app.services.parsers.chipmore_summary_parser import parse_and_save_chipmore_summary
        fallback_osat_name = "Chipmore" if str(osat_name).strip().lower() == "chipmore" else osat_name
        return parse_and_save_chipmore_summary(filepath, db, user_id, fallback_osat_name)
